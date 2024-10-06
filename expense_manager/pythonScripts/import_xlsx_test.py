import openpyxl
import argparse
import abc

class DataPack(abc.ABC):
    def __init__(self):
        pass

    @abc.abstractmethod
    def load_data(self):
        pass

    @abc.abstractmethod
    def display_data(self):
        pass

    @abc.abstractmethod
    def display_columns(self, columns):
        pass




class DataPackMono(DataPack):
    def __init__(self, file_name):
        super().__init__()
        self.file_name = file_name
        self.bank_name = "mono"
        self.headers = []
        self.headers_map = {}
        self.source_data = []

    def load_headers_map(self):
        short_headers = ['date', 'description', 'mcc', 'card_amount_uah', 'operation_amount',
                              'operation_currency', 'exchange_rate', 'commission_uah','cashback_uah','balance']
        for i in range(len(self.headers)):
            self.headers_map[self.headers[i]] = short_headers[i]
        


    def load_data(self):
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active

            # We read the headings from the first line
            self.headers = [cell.value for cell in sheet[1]]
            self.load_headers_map()

            # We read all the lines after the headings
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # We create a dictionary for each line of data
                row_data = {}
                for i, header in enumerate(self.headers):

                    # row_data[header] = row[excel_headers.index(header)]
                    row_data[self.headers_map[header]] = row[self.headers.index(header)]

                self.source_data.append(row_data)
        except FileNotFoundError as e:
            print(f"File '{self.file_name}' not found. {e}")
        except Exception as e:
            print(f"An error occurred while reading the file: {e}")

    def display_data(self):
        print(f"Bank statement from {self.bank_name}:")
        for row in self.source_data:
            print(row)


    def display_columns(self, columns):
        # We check whether all the transferred columns are in the headers
        invalid_columns = [col for col in columns if col not in self.headers]
        if invalid_columns:
            print(f"The following columns were not found: {', '.join(invalid_columns)}")
            return

        # We display column headings
        print(f"Column data: {', '.join(columns)}")

        # We output data for each line
        for row in self.source_data:
            row_values = [str(row.get(col, "No data available")) for col in columns]
            print(", ".join(row_values))



class DataPackPrivat(DataPack):
    def __init__(self, file_name):
        super().__init__()
        self.file_name = file_name
        self.bank_name = "privat"
        self.headers = []
        self.headers_map = {}
        self.source_data = []



    def load_headers_map(self):
        short_headers = ['date', 'time', 'category', 'card', 'description', 'card_amount', 'card_currency', 'transaction_amount',
                         'transaction_currency', 'balance_end', 'balance_currency', 'none']
        for i in range(len(self.headers)):
            self.headers_map[self.headers[i]] = short_headers[i]

    def load_data(self):
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active

            # We read the headings from the first line
            self.headers = [cell.value for cell in sheet[2]]
            self.load_headers_map()

            # We read all the lines after the headings
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # We create a dictionary for each line of data
                row_data = {}
                for i, header in enumerate(self.headers):

                    # row_data[header] = row[self.headers.index(header)]
                    row_data[self.headers_map[header]] = row[self.headers.index(header)]

                self.source_data.append(row_data)
        except FileNotFoundError as e:
            print(f"File '{self.file_name}' not found. {e}")
        except Exception as e:
            print(f"An error occurred while reading the file: {e}")

    def display_data(self):
        print(f"Bank statement from {self.bank_name}:")
        for row in self.source_data:
            print(row)

    def display_columns(self, columns):
        # We check whether all the transferred columns are in the headers
        invalid_columns = [col for col in columns if col not in self.headers]
        if invalid_columns:
            print(f"The following columns were not found: {', '.join(invalid_columns)}")
            return

        # We display column headings
        print(f"Column data: {', '.join(columns)}")

        # We output data for each line
        for row in self.source_data:
            row_values = [str(row.get(col, "No data available")) for col in columns]
            print(", ".join(row_values))


def main():
    
    parser = argparse.ArgumentParser(description="Gets file names and paths to them.")
    
    parser.add_argument('file1', type=str, help="The path to the first file")
    parser.add_argument('file2', type=str, help="The path to the first file")
    
    args = parser.parse_args()

    file1_path = args.file1
    file2_path = args.file2



    mono_data_pack = DataPackMono(file1_path)
    privat_data_pack = DataPackPrivat(file2_path)

    mono_data_pack.load_data()
    privat_data_pack.load_data()

    privat_data_pack.display_data()


if __name__ == "__main__":
    main()
    
#C:\Users\lucky\Desktop\Test_data2\mono_acc4_14-09-2024_19-01-2024.xlsx C:\Users\lucky\Desktop\Test_data2\privat24_acc4_22-12-2023_14-02-2024.xlsx
